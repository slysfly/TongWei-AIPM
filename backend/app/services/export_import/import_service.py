"""
通维AI项目管理系统 - 数据导入服务
支持 Excel、CSV 格式的任务导入
"""

import io
import csv
from datetime import datetime
from typing import List, Dict, Any, Tuple

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from app.models import Task, Project
from app.core.exceptions import NotFoundException, ValidationException
from app.services.export_import.export_service import ExportService


class ImportService:
    """数据导入服务"""

    # 字段映射配置
    FIELD_MAPPINGS = {
        "任务名称": "name",
        "name": "name",
        "任务": "name",
        "描述": "description",
        "description": "description",
        "状态": "status",
        "status": "status",
        "优先级": "priority",
        "priority": "priority",
        "负责人": "assignee",
        "assignee": "assignee",
        "负责人ID": "assignee_id",
        "assignee_id": "assignee_id",
        "截止日期": "planned_end",
        "planned_end": "planned_end",
        "结束日期": "planned_end",
        "预估工时": "estimated_hours",
        "estimated_hours": "estimated_hours",
        "实际工时": "actual_hours",
        "actual_hours": "actual_hours",
        "进度": "progress",
        "progress": "progress",
        "标签": "labels",
        "labels": "labels",
    }

    STATUS_REVERSE_MAP = {
        "待办": "backlog",
        "待开始": "todo",
        "进行中": "in_progress",
        "评审中": "in_review",
        "测试中": "testing",
        "已完成": "done",
        "已取消": "cancelled",
        "backlog": "backlog",
        "todo": "todo",
        "in_progress": "in_progress",
        "in_review": "in_review",
        "testing": "testing",
        "done": "done",
        "cancelled": "cancelled",
    }

    PRIORITY_REVERSE_MAP = {
        "P1-最高": 1,
        "P2-高": 2,
        "P3-中": 3,
        "P4-低": 4,
        "P5-最低": 5,
        "1": 1,
        "2": 2,
        "3": 3,
        "4": 4,
        "5": 5,
        "P1": 1,
        "P2": 2,
        "P3": 3,
        "P4": 4,
        "P5": 5,
    }

    @classmethod
    def parse_excel(cls, file_content: bytes) -> List[Dict[str, Any]]:
        """解析 Excel 文件"""
        try:
            import openpyxl
        except ImportError:
            raise ValidationException(message="openpyxl 未安装，无法解析 Excel")

        buffer = io.BytesIO(file_content)
        wb = openpyxl.load_workbook(buffer, data_only=True)
        ws = wb.active

        # 读取表头
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip() if cell.value else "")

        # 读取数据行
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or str(v).strip() == "" for v in row):
                continue
            row_dict = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    field_key = cls.FIELD_MAPPINGS.get(header, header)
                    row_dict[field_key] = row[idx]
            if row_dict.get("name"):
                data.append(row_dict)

        return data

    @classmethod
    def parse_csv(cls, file_content: bytes) -> List[Dict[str, Any]]:
        """解析 CSV 文件"""
        # 检测编码
        content_str = None
        for encoding in ["utf-8-sig", "utf-8", "gbk", "gb2312"]:
            try:
                content_str = file_content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if content_str is None:
            raise ValidationException(message="无法识别文件编码")

        buffer = io.StringIO(content_str)
        reader = csv.reader(buffer)

        # 读取表头
        try:
            headers = [h.strip() for h in next(reader)]
        except StopIteration:
            return []

        # 读取数据行
        data = []
        for row in reader:
            if not row or all(v.strip() == "" for v in row):
                continue
            row_dict = {}
            for idx, header in enumerate(headers):
                if header and idx < len(row):
                    field_key = cls.FIELD_MAPPINGS.get(header, header)
                    row_dict[field_key] = row[idx].strip()
            if row_dict.get("name"):
                data.append(row_dict)

        return data

    @classmethod
    def validate_task_data(cls, data: List[Dict[str, Any]]) -> Tuple[bool, List[str]]:
        """验证任务数据"""
        errors = []

        if not data:
            errors.append("数据为空，请检查文件内容")
            return False, errors

        for idx, row in enumerate(data, 1):
            # 验证任务名称
            name = row.get("name")
            if not name or str(name).strip() == "":
                errors.append(f"第 {idx} 行：任务名称不能为空")
                continue

            # 验证状态
            status = str(row.get("status", "")).strip()
            if status and status not in cls.STATUS_REVERSE_MAP:
                errors.append(f"第 {idx} 行：状态 '{status}' 无效")

            # 验证优先级
            priority = str(row.get("priority", "")).strip()
            if priority and priority not in cls.PRIORITY_REVERSE_MAP:
                errors.append(f"第 {idx} 行：优先级 '{priority}' 无效")

            # 验证进度
            progress = row.get("progress")
            if progress is not None and progress != "":
                try:
                    p = float(progress)
                    if p < 0 or p > 100:
                        errors.append(f"第 {idx} 行：进度必须在 0-100 之间")
                except (ValueError, TypeError):
                    errors.append(f"第 {idx} 行：进度格式无效")

            # 验证工时
            for field in ["estimated_hours", "actual_hours"]:
                value = row.get(field)
                if value is not None and value != "":
                    try:
                        float(value)
                    except (ValueError, TypeError):
                        errors.append(f"第 {idx} 行：{field} 格式无效")

        return len(errors) == 0, errors

    @classmethod
    def normalize_task_data(cls, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """规范化任务数据"""
        normalized = []
        for row in data:
            normalized_row = {}

            # 任务名称
            normalized_row["name"] = str(row.get("name", "")).strip()

            # 描述
            desc = row.get("description")
            normalized_row["description"] = str(desc).strip() if desc else None

            # 状态
            status = str(row.get("status", "todo")).strip()
            normalized_row["status"] = cls.STATUS_REVERSE_MAP.get(status, "todo")

            # 优先级
            priority = str(row.get("priority", "3")).strip()
            normalized_row["priority"] = cls.PRIORITY_REVERSE_MAP.get(priority, 3)

            # 负责人ID
            assignee_id = row.get("assignee_id")
            if assignee_id:
                normalized_row["assignee_id"] = str(assignee_id).strip() or None
            else:
                normalized_row["assignee_id"] = None

            # 截止日期
            planned_end = row.get("planned_end")
            if planned_end:
                try:
                    if isinstance(planned_end, datetime):
                        normalized_row["planned_end"] = planned_end
                    else:
                        from dateutil import parser
                        normalized_row["planned_end"] = parser.parse(str(planned_end))
                except Exception:
                    normalized_row["planned_end"] = None
            else:
                normalized_row["planned_end"] = None

            # 工时
            for field in ["estimated_hours", "actual_hours"]:
                value = row.get(field)
                if value is not None and value != "":
                    try:
                        normalized_row[field] = float(value)
                    except (ValueError, TypeError):
                        normalized_row[field] = 0
                else:
                    normalized_row[field] = 0

            # 进度
            progress = row.get("progress")
            if progress is not None and progress != "":
                try:
                    normalized_row["progress"] = float(progress)
                except (ValueError, TypeError):
                    normalized_row["progress"] = 0
            else:
                normalized_row["progress"] = 0

            # 标签
            labels = row.get("labels")
            if labels:
                if isinstance(labels, str):
                    normalized_row["labels"] = [l.strip() for l in labels.split(",") if l.strip()]
                elif isinstance(labels, list):
                    normalized_row["labels"] = labels
                else:
                    normalized_row["labels"] = []
            else:
                normalized_row["labels"] = []

            normalized.append(normalized_row)

        return normalized

    @classmethod
    async def import_tasks(
        cls,
        db: AsyncSession,
        project_id: str,
        data: List[Dict[str, Any]],
        user_id: str
    ) -> Dict[str, Any]:
        """批量导入任务"""
        # 验证项目存在
        project_result = await db.execute(
            select(Project).where(and_(Project.id == project_id, Project.is_deleted == False))
        )
        project = project_result.scalar_one_or_none()
        if not project:
            raise NotFoundException(message="项目不存在")

        # 规范化数据
        normalized_data = cls.normalize_task_data(data)

        # 获取当前项目下的任务数量（用于生成 WBS 编码）
        count_result = await db.execute(
            select(func.count(Task.id)).where(
                and_(Task.project_id == project_id, Task.is_deleted == False)
            )
        )
        existing_count = count_result.scalar() or 0

        # 批量创建任务
        created_count = 0
        failed_count = 0
        failed_details = []

        for idx, row in enumerate(normalized_data):
            try:
                existing_count += 1
                wbs_code = str(existing_count)

                task = Task(
                    project_id=project_id,
                    wbs_code=wbs_code,
                    name=row["name"],
                    description=row.get("description"),
                    status=row["status"],
                    priority=row["priority"],
                    assignee_id=row.get("assignee_id"),
                    planned_end=row.get("planned_end"),
                    estimated_hours=row.get("estimated_hours", 0),
                    actual_hours=row.get("actual_hours", 0),
                    progress=row.get("progress", 0),
                    labels=row.get("labels", []),
                    sort_order=existing_count,
                )
                db.add(task)
                created_count += 1
            except Exception as e:
                failed_count += 1
                failed_details.append(f"第 {idx + 1} 行 ({row.get('name', '未知')}): {str(e)}")

        await db.commit()

        return {
            "total": len(normalized_data),
            "created": created_count,
            "failed": failed_count,
            "failed_details": failed_details,
        }

    @classmethod
    def generate_template_excel(cls) -> bytes:
        """生成导入模板 Excel 文件"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ValidationException(message="openpyxl 未安装，无法生成模板")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "任务导入模板"

        # 标题样式
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        headers = [col["label"] for col in ExportService.TASK_EXPORT_COLUMNS]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入示例数据
        example_data = [
            ["示例任务1", "这是一个示例任务描述", "待开始", "P3-中", "", "2026-06-30", "8", "0", "0", "前端,紧急"],
            ["示例任务2", "另一个示例任务", "进行中", "P2-高", "", "2026-07-15", "16", "8", "50", "后端"],
        ]

        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # 添加说明sheet
        ws_info = wb.create_sheet("填写说明")
        info_data = [
            ["字段", "说明", "可选值"],
            ["任务名称", "必填，任务的名称", ""],
            ["描述", "可选，任务的详细描述", ""],
            ["状态", "可选，默认为待开始", "待办/待开始/进行中/评审中/测试中/已完成/已取消"],
            ["优先级", "可选，默认为P3-中", "P1-最高/P2-高/P3-中/P4-低/P5-最低"],
            ["负责人", "可选，填写用户ID或名称", ""],
            ["截止日期", "可选，格式YYYY-MM-DD", ""],
            ["预估工时", "可选，数字", ""],
            ["实际工时", "可选，数字", ""],
            ["进度", "可选，0-100的数字", ""],
            ["标签", "可选，多个标签用逗号分隔", ""],
        ]

        for row_idx, row_data in enumerate(info_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_info.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if row_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        ws_info.column_dimensions["A"].width = 15
        ws_info.column_dimensions["B"].width = 30
        ws_info.column_dimensions["C"].width = 50

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
