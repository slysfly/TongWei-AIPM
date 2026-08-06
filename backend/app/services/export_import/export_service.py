"""
通维AI项目管理系统 - 数据导出服务
支持 Excel、CSV、PDF 格式的任务导出
"""

import io
import csv
from datetime import datetime
from typing import List, Dict, Any, Tuple, Optional

from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from sqlalchemy.orm import selectinload

from app.models import Task, Project, User, Risk
from app.core.exceptions import NotFoundException, ValidationException


class ExportService:
    """数据导出服务"""

    # 任务导出字段配置
    TASK_EXPORT_COLUMNS = [
        {"key": "name", "label": "任务名称", "width": 30},
        {"key": "description", "label": "描述", "width": 40},
        {"key": "status", "label": "状态", "width": 12},
        {"key": "priority", "label": "优先级", "width": 10},
        {"key": "assignee", "label": "负责人", "width": 15},
        {"key": "planned_end", "label": "截止日期", "width": 15},
        {"key": "estimated_hours", "label": "预估工时", "width": 12},
        {"key": "actual_hours", "label": "实际工时", "width": 12},
        {"key": "progress", "label": "进度(%)", "width": 10},
        {"key": "labels", "label": "标签", "width": 20},
    ]

    STATUS_LABEL_MAP = {
        "backlog": "待办",
        "todo": "待开始",
        "in_progress": "进行中",
        "in_review": "评审中",
        "testing": "测试中",
        "done": "已完成",
        "cancelled": "已取消",
    }

    PRIORITY_LABEL_MAP = {
        1: "P1-最高",
        2: "P2-高",
        3: "P3-中",
        4: "P4-低",
        5: "P5-最低",
    }

    @classmethod
    async def _fetch_tasks(
        cls,
        db: AsyncSession,
        project_id: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> List[Task]:
        """获取任务列表"""
        query = select(Task).where(
            and_(Task.project_id == project_id, Task.is_deleted == False)
        ).options(selectinload(Task.assignee))

        if filters:
            if filters.get("status"):
                query = query.where(Task.status == filters["status"])
            if filters.get("priority"):
                query = query.where(Task.priority == filters["priority"])
            if filters.get("assignee_id"):
                query = query.where(Task.assignee_id == filters["assignee_id"])

        query = query.order_by(Task.sort_order, Task.created_at)
        result = await db.execute(query)
        return result.scalars().all()

    @classmethod
    def _task_to_row(cls, task: Task) -> Dict[str, Any]:
        """将任务转换为导出行数据"""
        return {
            "name": task.name or "",
            "description": task.description or "",
            "status": cls.STATUS_LABEL_MAP.get(task.status, task.status),
            "priority": cls.PRIORITY_LABEL_MAP.get(task.priority, str(task.priority)),
            "assignee": task.assignee.full_name if task.assignee else "",
            "planned_end": task.planned_end.strftime("%Y-%m-%d") if task.planned_end else "",
            "estimated_hours": float(task.estimated_hours) if task.estimated_hours else 0,
            "actual_hours": float(task.actual_hours) if task.actual_hours else 0,
            "progress": float(task.progress) if task.progress else 0,
            "labels": ", ".join(task.labels) if task.labels else "",
        }

    @classmethod
    async def export_tasks_to_excel(
        cls,
        db: AsyncSession,
        project_id: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """导出任务为 Excel 格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise ValidationException(message="openpyxl 未安装，无法导出 Excel")

        tasks = await cls._fetch_tasks(db, project_id, filters)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "任务列表"

        # 标题样式
        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1890FF", end_color="1890FF", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 写入表头
        for col_idx, col in enumerate(cls.TASK_EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col["label"])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = col["width"]

        # 写入数据
        for row_idx, task in enumerate(tasks, 2):
            row_data = cls._task_to_row(task)
            for col_idx, col in enumerate(cls.TASK_EXPORT_COLUMNS, 1):
                value = row_data.get(col["key"], "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动调整行高
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 20

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    @classmethod
    async def export_tasks_to_csv(
        cls,
        db: AsyncSession,
        project_id: str,
        filters: Optional[Dict[str, Any]] = None
    ) -> bytes:
        """导出任务为 CSV 格式"""
        tasks = await cls._fetch_tasks(db, project_id, filters)

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # 写入 BOM 和表头
        headers = [col["label"] for col in cls.TASK_EXPORT_COLUMNS]
        writer.writerow(headers)

        # 写入数据
        for task in tasks:
            row_data = cls._task_to_row(task)
            row = [row_data.get(col["key"], "") for col in cls.TASK_EXPORT_COLUMNS]
            writer.writerow(row)

        return buffer.getvalue().encode("utf-8-sig")

    @classmethod
    async def export_project_report_to_pdf(
        cls,
        db: AsyncSession,
        project_id: str
    ) -> Tuple[bytes, str]:
        """
        导出项目报告为真实 PDF 文件（基于 reportlab）。
        若未安装 reportlab，抛出明确异常，避免对外呈现"生成成功"的假象。
        """
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.lib.units import cm
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            )
            from reportlab.pdfbase import pdfmetrics
            from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        except ImportError:
            raise ValidationException(message="reportlab 未安装，无法导出 PDF")

        # 注册中文字体（reportlab 内置 CID 字体，无需额外字体文件）
        try:
            pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
            cjk_font = "STSong-Light"
        except Exception:
            cjk_font = "Helvetica"

        # 获取项目信息
        project_result = await db.execute(
            select(Project).where(and_(Project.id == project_id, Project.is_deleted == False))
        )
        project = project_result.scalar_one_or_none()
        if not project:
            raise NotFoundException(message="项目不存在")

        # 获取任务统计
        tasks = await cls._fetch_tasks(db, project_id)
        total_tasks = len(tasks)
        completed_tasks = sum(1 for t in tasks if t.status == "done")
        in_progress_tasks = sum(1 for t in tasks if t.status == "in_progress")
        overdue_tasks = sum(
            1 for t in tasks
            if t.status != "done" and t.planned_end and t.planned_end < datetime.now()
        )

        # 获取风险信息
        risk_result = await db.execute(
            select(Risk).where(Risk.project_id == project_id)
        )
        risks = risk_result.scalars().all()
        high_risks = [r for r in risks if r.risk_score and float(r.risk_score) > 0.6]

        # 构建 PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            leftMargin=2 * cm, rightMargin=2 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
            title=f"项目报告 - {project.name}",
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "TitleCJK", parent=styles["Title"], fontName=cjk_font, textColor=colors.HexColor("#1890FF")
        )
        h2_style = ParagraphStyle(
            "H2CJK", parent=styles["Heading2"], fontName=cjk_font, textColor=colors.HexColor("#333333")
        )
        normal_style = ParagraphStyle(
            "NormalCJK", parent=styles["Normal"], fontName=cjk_font, fontSize=10, leading=14
        )
        cell_style = ParagraphStyle(
            "CellCJK", parent=styles["Normal"], fontName=cjk_font, fontSize=9, leading=12
        )

        story = []
        story.append(Paragraph(f"项目报告：{project.name}", title_style))
        story.append(Paragraph(
            f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style
        ))
        story.append(Spacer(1, 0.4 * cm))

        # 项目概览
        story.append(Paragraph("一、项目概览", h2_style))
        overview_rows = [
            ["项目描述", project.description or "暂无描述"],
            ["项目状态", str(project.status)],
            ["开始日期", str(project.start_date) if project.start_date else "未设置"],
            ["结束日期", str(project.end_date) if project.end_date else "未设置"],
        ]
        overview_tbl = Table(
            [[Paragraph(k, cell_style), Paragraph(str(v), cell_style)] for k, v in overview_rows],
            colWidths=[4 * cm, 12 * cm],
        )
        overview_tbl.setStyle(TableStyle([
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e8e8e8")),
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#f5f5f5")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(overview_tbl)
        story.append(Spacer(1, 0.4 * cm))

        # 任务统计卡片
        story.append(Paragraph("二、任务统计", h2_style))
        stat_header = ["总任务数", "已完成", "进行中", "已逾期"]
        stat_values = [total_tasks, completed_tasks, in_progress_tasks, overdue_tasks]
        stat_tbl = Table(
            [stat_header, [str(v) for v in stat_values]],
            colWidths=[4 * cm] * 4,
        )
        stat_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1890FF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), cjk_font),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e8e8e8")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(stat_tbl)
        story.append(Spacer(1, 0.4 * cm))

        # 任务列表
        story.append(Paragraph("三、任务列表", h2_style))
        task_header = ["任务名称", "状态", "优先级", "负责人", "进度", "预估/实际工时"]
        task_rows = [task_header]
        for task in tasks:
            status_label = cls.STATUS_LABEL_MAP.get(task.status, task.status)
            priority_label = cls.PRIORITY_LABEL_MAP.get(task.priority, str(task.priority))
            assignee_name = task.assignee.full_name if task.assignee else "未分配"
            progress = float(task.progress) if task.progress else 0
            est_hours = float(task.estimated_hours) if task.estimated_hours else 0
            act_hours = float(task.actual_hours) if task.actual_hours else 0
            task_rows.append([
                Paragraph(task.name, cell_style),
                Paragraph(status_label, cell_style),
                Paragraph(priority_label, cell_style),
                Paragraph(assignee_name, cell_style),
                Paragraph(f"{progress:.0f}%", cell_style),
                Paragraph(f"{est_hours:.1f}h / {act_hours:.1f}h", cell_style),
            ])
        task_tbl = Table(task_rows, colWidths=[4.5 * cm, 2 * cm, 2 * cm, 2.5 * cm, 2 * cm, 3 * cm], repeatRows=1)
        task_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1890FF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, -1), cjk_font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e8e8e8")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafafa")]),
        ]))
        story.append(task_tbl)
        story.append(Spacer(1, 0.4 * cm))

        # 风险分析
        story.append(Paragraph("四、风险分析", h2_style))
        if high_risks:
            story.append(Paragraph(f"发现 {len(high_risks)} 个高风险项，需要重点关注。", normal_style))
            risk_header = ["风险名称", "类别", "概率", "影响", "风险评分", "应对策略"]
            risk_rows = [risk_header]
            for risk in high_risks:
                risk_score = float(risk.risk_score) if risk.risk_score else 0
                risk_rows.append([
                    Paragraph(risk.name, cell_style),
                    Paragraph(str(risk.category or ""), cell_style),
                    Paragraph(f"{float(risk.probability):.0%}" if risk.probability else "0%", cell_style),
                    Paragraph(f"{float(risk.impact):.0%}" if risk.impact else "0%", cell_style),
                    Paragraph(f"{risk_score:.2f}", cell_style),
                    Paragraph(risk.response_strategy or "未制定", cell_style),
                ])
            risk_tbl = Table(
                risk_rows,
                colWidths=[3 * cm, 2 * cm, 1.8 * cm, 1.8 * cm, 2 * cm, 4.4 * cm],
                repeatRows=1,
            )
            risk_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5222d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), cjk_font),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e8e8e8")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            story.append(risk_tbl)
        else:
            story.append(Paragraph("暂无高风险项。", normal_style))

        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue(), "pdf"

    @classmethod
    async def export_gantt_to_image(
        cls,
        db: AsyncSession,
        project_id: str
    ) -> Optional[bytes]:
        """
        导出甘特图为 PNG 图片
        由于需要浏览器渲染环境，此处返回 None，建议前端使用 html2canvas 实现
        """
        return None
